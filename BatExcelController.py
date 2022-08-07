from BatDesigner import BatDesigner

import pandas as pd

from os.path import exists

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Alignment, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import DataBarRule

from abc import ABC

from itertools import islice


# Abstract
class BatExcelController(ABC):
    def __init__(self):
        self.__color_index = 0
        self.bat_csv = pd.DataFrame
        self.all_bats = list()
        self.path = ""
        self.ten_minute_intervals = list()
        self.ten_minute_classes = list()
        self.designer = BatDesigner()

    def __color_wheel(self):
        # colours = ["00ffefd4", "00fffeef", "00eff0ff"]
        colours = self.designer.color_wheel_classes
        pick = colours[self.__color_index]
        self.__color_index = (self.__color_index + 1) % len(colours)
        return pick

    def import_excel(self, path, one_file=False):
        workbook = load_workbook(path)
        bat_columns = dict()
        # Get workbook active sheet object
        # from the active attribute
        
        if not one_file:
            
            worksheet = workbook.active
            table = worksheet.tables["Table1"]
            
            """
            # iterate through header
            for i in range(1, worksheet.max_column):
                cell_obj = worksheet.cell(row=1, column=i)
                value = cell_obj.value
                if value in self.all_bats:
                    bat_columns[value] = cell_obj.coordinate
            """

            data = worksheet.values
            cols = next(data)[1:]
            data = list(data)
            idx = [r[0] for r in data]
            data = (islice(r, 1, None) for r in data)
            df = pd.DataFrame(data, index=idx, columns=cols)

            # rename to internal representation
            df.rename(columns=self.designer.reversed_columns(), inplace=True)

            # is it identical?
            #diff_manual_id = min(df['manual_id'].isin(self.bat_csv['manual_id']).values)
            #diff_reference_id = min(df['reference_id'].isin(self.bat_csv['reference_id']).values)
            equal_datasets = min(df['file_name'].isin(self.bat_csv['file_name']).values)  # bool

            if equal_datasets:
                #self.bat_csv["manual_id"] = df["manual_id"].copy()
                #self.bat_csv["reference_id"] = df["reference_id"].copy
                print("Datasets are equal! Merging in new file")

                # id column with matching content
                id = 'file_name'
                cols_to_replace = ['manual_id', 'reference_id']
                self.bat_csv.loc[self.bat_csv[id].isin(df[id]), cols_to_replace] = df.loc[df[id].isin(self.bat_csv[id]), cols_to_replace].values

                file = path.split("/")[-1]
                file = file.split(".")[0]
                self.export_excel(file, skip=True)
            
        else:
            

        

    def export_excel(self, output_file, skip=False):
        # dataframe to spreadsheet

        # Skip, if file already exists
        save_path = f"{self.path}/{output_file}.xlsx"
        if not skip and exists(save_path):
            self.import_excel(path=save_path)
            print("File already exiting!")
            return

        # create a workbook and grab active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        # worksheet.title = self.location

        # loop over dataframe values and append to worksheet
        for i in dataframe_to_rows(self.bat_csv, index=False, header=True):
            worksheet.append(i)

        # Create a Excel Table

        table = Table(displayName="Table1",
                      ref="A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style

        worksheet.add_table(table)

        # Freeze first Row
        worksheet.freeze_panes = 'A2'

        # Adjust columns
        adjust_columns = {
            "file_path": 30,
            "file_name": 40,
            "date": 12,
            "time": 12,
            "autofilled": 30,
            "manual_id": 30,
            "global_10_min_class": 20,
            "MbraMmys": 15,
            "reference_id": 15,
        }

        # Adjust all 10min classes
        for bat in self.all_bats:
            adjust_columns[self.designer.bat_class(bat)] = 14

        # col = worksheet.column_dimensions['A']
        # col.font = Font(bold=True)

        # Data Bars
        rule = DataBarRule(start_type='num',
                           start_value=0,
                           end_type='num',
                           end_value=1,
                           color=colors.BLUE)
                           # color="638EC6")

        center_columns = ["file_name", "manual_id", "date", "time", "autofilled", "global_10_min_class"]

        # =IF(COUNTIF(B5:B12,"*Ppip*"),1,"")

        letter_manual_id = ""

        # get letter for manual ID
        for i in range(1, worksheet.max_column + 1):  # Header line
            cell = worksheet.cell(row=1, column=i)
            value = cell.value
            if value == "manual_id":
                letter_manual_id = get_column_letter(i)

        colour = self.__color_wheel()

        fill = PatternFill(start_color=colour,
                           end_color=colour,
                           fill_type='solid')

        border_color = "00D3D3D3"  # Light gray
        # border_color = "00565656" # dark gray
        thin_border = Border(left=Side(style='thin', color=border_color),
                             right=Side(style='thin', color=border_color),
                             top=Side(style='thin', color=border_color),
                             bottom=Side(style='thin', color=border_color),
                             )

        for i in range(1, worksheet.max_column + 1):  # Header line
            cell = worksheet.cell(row=1, column=i)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
            value = cell.value

            # 10 min classes
            last_class_row = worksheet.max_row + 1
            column_letter = get_column_letter(i)

            # paint rows
            if value == "global_10_min_class":
                for row in range(2, worksheet.max_row + 1):

                    if row - 2 in self.ten_minute_intervals:
                        colour = self.__color_wheel()
                        fill = PatternFill(start_color=colour,
                                           end_color=colour,
                                           fill_type='solid'
                                           )
                    for column in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row, column=column)
                        cell.fill = fill
                        cell.border = thin_border
            # 10 min classes formula
            if value in self.ten_minute_classes:

                for row in range(worksheet.max_row + 1, 1, -1):

                    if row - 2 in self.ten_minute_intervals and row > 1:
                        text = value.split("_")
                        bat = text[0]
                        row_start = last_class_row - 1
                        row_end = row
                        value_range = f"${letter_manual_id}${row_start}:${letter_manual_id}${row}"
                        last_class_row = row
                        cell = worksheet.cell(row=row, column=i)
                        cell.value = f'=IF(COUNTIF({value_range}, "*{bat}*"), 1, "")'

            # Apply Data Bars
            if value in self.all_bats:
                start_cell = worksheet.cell(row=1, column=i)
                end_cell = worksheet.cell(row=worksheet.max_row, column=i)
                worksheet.conditional_formatting.add(f"{start_cell.coordinate}:{end_cell.coordinate}", rule)

                # Align bat numbers left
                for row in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=i)
                    cell.alignment = Alignment(horizontal='left')

            # Adjust Columns
            if value in adjust_columns.keys():
                worksheet.column_dimensions[get_column_letter(i)].width = adjust_columns[value]

            if value in center_columns:

                for row in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=i)
                    cell.alignment = Alignment(horizontal='center')

                if value == "manual_id":
                    col = worksheet.column_dimensions[get_column_letter(i)]
                    col.font = Font(bold=True)

        # Finally, rewrite Header
        for i in range(1, worksheet.max_column + 1):  # Header line
            cell = worksheet.cell(row=1, column=i)
            value = cell.value

            cell.value = self.designer.get_column(value)


        # save workbook
        workbook.save(save_path)


