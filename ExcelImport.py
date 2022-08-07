from typing import Dict, List, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Font, Fill, Alignment, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import DataBarRule

import pandas as pd

from Standort import Standort
from Aufnahme import Aufnahme
#from DataAnalysis import Projekt

class ExcelImport:
    def __init__(self, path:str) -> None:
        self.import_excel(path=path)
        
    def __split_excel_string(self, text:str):
        range_start, range_end = text.split(":")
        min_col, min_row = coordinate_from_string(range_start)
        max_col, max_row = coordinate_from_string(range_end)
        
        return (min_col, max_col, min_row, max_row)
                
    
    def import_excel(self, path:str):
        workbook = load_workbook(path)
        sheet_names = workbook.sheetnames
        
        standorte = dict()
        temperaturen = dict()
        
        
        for sheet in sheet_names:
            if "Standort " in sheet:
                worksheet = workbook[sheet]
                table_name, table_range = worksheet.tables.items()[0]               
                min_col, max_col, min_row, max_row = self.__split_excel_string(table_range)
            
                data = pd.read_excel(path, sheet_name=sheet, nrows=max_row-1, usecols=f"{min_col}:{max_col}")
            
                standort_name, aufnahme_date = sheet.split(" - ")                
                aufnahme = Aufnahme(name=aufnahme_date)
                aufnahme.add_data(data)
                                
                if standort_name not in standorte:               
                    standort = Standort(name=standort_name)
                    standorte[standort_name] = standort
                    
                else:
                    standort = standorte[standort_name]
                
                standort.add_aufnahme(aufnahme)
                
            if "Temperaturdaten" in sheet:
                worksheet = workbook[sheet]
                for table_name, table_range in worksheet.tables.items():
                    min_col, max_col, min_row, max_row = self.__split_excel_string(table_range)
                    
                    standort_aufnahme = worksheet[f"{min_col}1"].value
                    data = pd.read_excel(path, sheet_name=sheet, nrows=max_row, skiprows=min_row - 1, usecols=f"{min_col}:{max_col}")
                    # print(table_name, table_range)
                    
                    #strange workaround to name Temperatur equal
                    for column in data.columns:
                        if "Temperatur" in column:
                            data = data.rename(columns={column: 'Temperatur_(°C)'})
                    
                    
                    
                    # mean duplicated dates
                    data = data.groupby('Temperatur_(°C)', as_index=False).mean()
                    
                    temperaturen[standort_aufnahme] = data
                    
                    
        for standort_aufnahme in temperaturen:
            standort_name, aufnahme_name = standort_aufnahme.split(" - ")
            
            standort = standorte[standort_name]
            aufnahme = standort.get_aufnahme(aufnahme_name)
            #aufnahme.da
             
        
        return standorte
                
                